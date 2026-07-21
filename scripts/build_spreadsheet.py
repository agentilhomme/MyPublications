#!/usr/bin/env python3
"""Build a skimmable spreadsheet of publications from data/publications.json.

Usage:
    python -X utf8 scripts/build_spreadsheet.py

Run from the repository root. Reads data/publications.json (see
fetch_publications.py) and writes publications.xlsx with one row per paper:
title, year, venue, type, and co-authors.
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT / "data" / "publications.json"
OUT_PATH = ROOT / "publications.xlsx"

FONT_NAME = "Arial"


def strip_html(text):
    return re.sub(r"<[^>]+>", "", text)


def main():
    data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    works = sorted(data["works"], key=lambda w: (w["year"] or 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    headers = ["Title", "Year", "Venue", "Type", "Co-authors"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for w in works:
        title = strip_html(w["title"])
        authors = ", ".join(w.get("authors", []))
        ws.append([title, w["year"], w["venue"], w["type"], authors])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 55, "B": 8, "C": 28, "D": 12, "E": 45}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Saved {len(works)} rows to {OUT_PATH}")


if __name__ == "__main__":
    main()
